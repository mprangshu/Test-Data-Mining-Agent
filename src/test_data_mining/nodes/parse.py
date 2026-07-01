"""
parse.py — Extract required fields + scenario types from the PRIMARY inputs.

Type: deterministic. Reads ``<input_path>/test_cases/`` (or ``input_path`` itself) and supports
``.xlsx``, ``.csv``, ``.json``, ``.txt`` (Gherkin / acceptance criteria). Produces
``parsed_fields`` (the list of fields to generate data for). Never crashes → notes to ``gaps``.

Convention (pivot §10): for tabular inputs, **column headers are field names**; columns that
*label* a scenario (``scenario_type`` / ``data_category`` / ``scenario_tag``) and id columns
(``test_case_id`` …) are not fields. For Gherkin ``.txt``, ``<placeholders>`` are the fields.
"""
from __future__ import annotations

import csv
import json
import os
import re
from collections import OrderedDict

from ..state import AgentState, ParsedField

# Columns that label/identify rather than name a data field.
SCENARIO_COLS = {"scenario_type", "data_category"}
TAG_COLS = {"scenario_tag"}
ID_COLS = {"test_case_id", "test_id", "testcase_id"}
_NON_FIELD = SCENARIO_COLS | TAG_COLS | ID_COLS

_CANON_SCEN = {"valid", "boundary", "negative", "edge"}

# (name substrings, category, constraints)
_CATEGORY_RULES = [
    (("email",), "Identity", ["required", "email_format"]),
    (("card_number", "card", "cvv", "iban"), "PII", ["required", "masked"]),
    (("currency",), "Reference", ["required", "ISO-4217"]),
    (("country",), "Reference", ["required", "ISO-3166"]),
    (("order_total", "total", "amount", "price", "balance"), "Financial", ["required", ">=0"]),
    (("item_count", "count", "quantity", "qty"), "Quantity", ["required", "integer"]),
    (("payment", "method"), "Reference", ["required"]),
    (("customer", "name", "first_name", "last_name"), "PII", ["required"]),
    (("created_at", "date", "timestamp", "time"), "Temporal", ["required", "ISO-8601"]),
    (("status",), "Reference", ["required"]),
    (("coupon", "code", "promo"), "Reference", []),
    (("order_id", "id"), "Identifier", ["required", "unique"]),
]


def _infer(name: str) -> tuple[str, list[str]]:
    # Infer a field category and constraint list from its name.
    # Output example: ("Identity", ["required", "email_format"]) for "user_email".
    low = name.lower()
    for subs, cat, cons in _CATEGORY_RULES:
        if any(s in low for s in subs):
            return cat, list(cons)
    return "General", ["required"]


def _norm_scenarios(values) -> list[str]:
    # Normalize raw scenario labels to the canonical set.
    # Example input: ["Valid", "NEGATIVE", "invalid"] -> ["valid", "negative"]
    out: list[str] = []
    for v in values:
        if v is None:
            continue
        key = str(v).strip().lower()
        if key in _CANON_SCEN and key not in out:
            out.append(key)
    return out


def _emit(acc: "OrderedDict[str, dict]", name: str, scenarios: list[str], ids: list[str]) -> None:
    # Add or update a parsed field entry in the accumulator.
    # This is called from _from_table/_from_txt to collect field metadata.
    # The accumulator entries look like:
    # {"email": {"category": "Identity", "constraints": [...], "scenarios": ["valid"], "ids": ["TC1"]}}
    name = (name or "").strip()
    if not name or name in _NON_FIELD:
        return
    if name not in acc:
        cat, cons = _infer(name)
        acc[name] = {"category": cat, "constraints": cons, "scenarios": [], "ids": []}
    for s in scenarios:
        if s not in acc[name]["scenarios"]:
            acc[name]["scenarios"].append(s)
    for i in ids:
        if i and i not in acc[name]["ids"]:
            acc[name]["ids"].append(i)


def _from_table(acc, headers: list[str], rows: list[dict]) -> None:
    # Extract candidate field names and labels from tabular inputs.
    # Fields are headers excluding scenario/ID columns.
    # Example: headers ["email","scenario_type","test_case_id"] -> fields ["email"].
    fields = [h for h in headers if h and h not in _NON_FIELD]
    scen_vals, id_vals = [], []
    for r in rows:
        for c in SCENARIO_COLS:
            if c in r:
                scen_vals.append(r[c])
        for c in ID_COLS:
            if r.get(c):
                id_vals.append(str(r[c]))
    scenarios = _norm_scenarios(scen_vals) or ["valid"]
    for h in fields:
        _emit(acc, h, scenarios, id_vals)


def _read_csv(path: str):
    # Read a CSV and return (headers, rows).
    # Sample output: (["email","order_total"], [{"email":"a@b.com","order_total":"12.34"}])
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def _read_xlsx(path: str):
    # Read an XLSX workbook, return (headers, rows) from the active sheet.
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(c).strip() if c is not None else "" for c in next(it, [])]
        rows = [
            {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
            for r in it
        ]
    finally:
        wb.close()
    return headers, rows


def _read_json(path: str):
    # Read JSON as either a list of records or a single object.
    # Sample output: (["email"], [{"email":"a@b.com"}])
    with open(path, encoding="utf-8") as f:
        doc = json.load(f)
    if isinstance(doc, list) and doc and isinstance(doc[0], dict):
        headers: list[str] = []
        for row in doc:
            for k in row:
                if k not in headers:
                    headers.append(k)
        return headers, doc
    if isinstance(doc, dict):
        if isinstance(doc.get("fields"), list):
            return [str(x) for x in doc["fields"]], []
        return list(doc.keys()), [doc]
    return [], []


def _select_primary(tables: list[tuple[list[str], list[dict]]]) -> tuple[list[str], list[dict]]:
    """Pick the original-rows table (most rows wins; merge any sharing identical headers).

    These rows pass through ``synthesise`` UNCHANGED (invariant #7, additive). Values are kept
    verbatim; only missing columns are filled with "" so every row carries the full column set.
    Schema-only inputs (Gherkin ``.txt``, a JSON ``{"fields": [...]}`` doc) contribute no rows.
    """
    # Choose the table with the most rows as the primary original dataset.
    # This output is used directly by synthesise to preserve verbatim rows.
    tables = [(h, r) for h, r in tables if h and r]
    if not tables:
        return [], []
    primary_headers = max(tables, key=lambda t: len(t[1]))[0]
    rows: list[dict] = []
    for headers, rs in tables:
        if headers == primary_headers:
            rows.extend(rs)
    norm = [{c: r.get(c, "") for c in primary_headers} for r in rows]
    return list(primary_headers), norm


def _from_txt(acc, text: str, gaps: list[str], src: str) -> None:
    # Extract placeholder names from Gherkin-style text and infer scenario tags.
    # Example text: "Given <email> and <order_total>" -> fields ["email","order_total"]
    placeholders = re.findall(r"<([^>]+)>", text)
    low = text.lower()
    scenarios = ["valid"]
    if any(w in low for w in ("invalid", "missing", "negative", "error", "reject", "not allowed")):
        scenarios.append("negative")
    if any(w in low for w in ("boundary", "maximum", "minimum", "zero", "limit", "empty")):
        scenarios.append("boundary")
    if "edge" in low:
        scenarios.append("edge")
    for p in placeholders:
        _emit(acc, p.strip(), scenarios, [])
    if not placeholders:
        gaps.append(f"parse: no <placeholders> found in {src}")


def parse(state: AgentState) -> dict:
    """LangGraph node: read the primary test-case inputs → parsed_fields."""
    base = state["input_path"]
    tc_dir = os.path.join(base, "test_cases")
    root = tc_dir if os.path.isdir(tc_dir) else base

    if not os.path.isdir(root):
        return {"parsed_fields": [], "gaps": [f"parse: input dir not found: {root}"]}

    acc: "OrderedDict[str, dict]" = OrderedDict()
    gaps: list[str] = []
    tables: list[tuple[list[str], list[dict]]] = []   # tabular files → original rows (verbatim)
    files = [f for f in sorted(os.listdir(root)) if os.path.isfile(os.path.join(root, f))]
    for fn in files:
        ext = os.path.splitext(fn)[1].lower()
        fp = os.path.join(root, fn)
        try:
            if ext == ".csv":
                headers, rows = _read_csv(fp)
                _from_table(acc, headers, rows); tables.append((headers, rows))
            elif ext == ".xlsx":
                headers, rows = _read_xlsx(fp)
                _from_table(acc, headers, rows); tables.append((headers, rows))
            elif ext == ".json":
                headers, rows = _read_json(fp)
                _from_table(acc, headers, rows); tables.append((headers, rows))
            elif ext == ".txt":
                with open(fp, encoding="utf-8") as f:
                    _from_txt(acc, f.read(), gaps, fn)
        except Exception as exc:  # never crash — flag and continue (spec §1.4)
            gaps.append(f"parse: skipped {fn} ({type(exc).__name__}: {exc})")

    # Convert the accumulated field metadata to TypedDict dataclasses.
    fields = [
        ParsedField(name=n, category=d["category"], constraints=d["constraints"],
                    source_test_ids=d["ids"], scenario_types=d["scenarios"] or ["valid"])
        for n, d in acc.items()
    ]
    if not fields:
        gaps.append("parse: no fields extracted from primary inputs")

    # Original rows pass through synthesise UNCHANGED (invariants #7/#9): the output keeps these
    # exact columns and starts from these exact rows, then appends generated rows.
    input_columns, input_rows = _select_primary(tables)
    if not input_rows:
        gaps.append("parse: no tabular rows found — output will be generated-only (still > 0 rows)")

    print(f"NODE_EXIT parse: {len(fields)} fields, {len(input_rows)} original rows "
          f"({len(input_columns)} cols) from {len(files)} file(s)")
    # Sample return structure:
    # {
    #   "parsed_fields": [ParsedField(...)],
    #   "input_rows": [{...}],
    #   "input_columns": ["email","order_total"],
    #   "input_row_count": 3,
    #   "gaps": []
    # }
    return {
        "parsed_fields": fields,
        "input_rows": input_rows,
        "input_columns": input_columns,
        "input_row_count": len(input_rows),
        "gaps": gaps,
    }
